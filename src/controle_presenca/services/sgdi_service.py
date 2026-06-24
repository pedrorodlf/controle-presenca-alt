import os
import secrets
from typing import List
from sqlalchemy.orm import Session
from ..database.models import Candidato, Aluno, HistoricoStatusCandidato, PontuacaoQuestionario
from ..utils.score_calculator import ScoreCalculator
from ..utils.google_drive import GoogleDriveDownloader

class SGDiService:
    def __init__(self, db: Session):
        self.db = db

    def cadastrar_candidato(self, nome: str, cpf: str, email: str, respostas: dict = None, perguntas_texto: dict = None):
        """
        Cadastra um novo candidato limpando dados (CPF e Nome) e evitando duplicatas.
        Opcionalmente calcula a pontuação socioeconômica real baseada no questionário,
        e salva as respostas individuais do questionário na tabela pontuacoes_questionario.
        """
        cpf_limpo = ''.join(c for c in cpf if c.isdigit())
        if len(cpf_limpo) != 11:
            return False, "❌ CPF inválido. Deve conter 11 dígitos."

        nome_limpo = nome.strip().upper()

        # Evita duplicidade por CPF
        c_existente = self.db.query(Candidato).filter(Candidato.cpf == cpf_limpo).first()
        if c_existente:
            return False, f"❌ Candidato com CPF {cpf_limpo} já cadastrado."

        # Calcula a pontuação real caso as respostas do formulário tenham sido informadas
        pontos = 0.0
        detalhes = {}
        if respostas:
            detalhes = ScoreCalculator.calcular_detalhes(respostas)
            pontos = sum(d["pontos"] for d in detalhes.values())

        novo_cand = Candidato(
            nome=nome_limpo,
            cpf=cpf_limpo,
            email=email.strip(),
            status='pendente',
            pontuacao_socioeconomica=pontos
        )
        self.db.add(novo_cand)
        self.db.flush()

        # Salva as respostas detalhadas do questionário no banco se perguntas_texto for informado
        if respostas and perguntas_texto:
            for q_num, det in detalhes.items():
                q_texto = perguntas_texto.get(q_num, f"Questão {q_num}")
                pont_quest = PontuacaoQuestionario(
                    candidato_id=novo_cand.id,
                    questao=str(q_texto)[:255] if q_texto else f"Questão {q_num}",
                    resposta=str(det["resposta"])[:500] if det["resposta"] else "",
                    pergunta_numero=int(q_num),
                    pontos=float(det["pontos"])
                )
                self.db.add(pont_quest)

        # Registrar no histórico
        historico = HistoricoStatusCandidato(
            candidato_id=novo_cand.id,
            candidato_nome=novo_cand.nome,
            candidato_cpf=novo_cand.cpf,
            status_anterior=None,
            status_novo='pendente',
            observacao="Cadastro inicial do candidato."
        )
        self.db.add(historico)
        self.db.commit()
        return True, f"✅ Candidato {nome_limpo} cadastrado com sucesso!"

    def gerar_ranking(self, limite: int = 60):
        """Busca candidatos pendentes ordenados por pontuação decrescente"""
        return self.db.query(Candidato).filter(
            Candidato.status == 'pendente'
        ).order_by(Candidato.pontuacao_socioeconomica.desc()).limit(limite).all()

    def aprovar_corte(self, quantidade: int):
        """
        Aprova candidatos do topo do ranking, respeitando o limite rígido de 60 discentes ativos.
        """
        total_ativos = self.db.query(Aluno).filter(Aluno.status == 'ATIVADO').count()
        if total_ativos + quantidade > 60:
            vagas = max(0, 60 - total_ativos)
            raise ValueError(
                f"⚠️ O corte de {quantidade} excede o limite máximo de 60 discentes ativos! "
                f"(Atuais: {total_ativos}, Vagas restantes: {vagas})"
            )

        aprovados = self.gerar_ranking(quantidade)
        for cand in aprovados:
            status_anterior = cand.status
            cand.status = 'aprovado'
            
            # Registrar no histórico
            historico = HistoricoStatusCandidato(
                candidato_id=cand.id,
                candidato_nome=cand.nome,
                candidato_cpf=cand.cpf,
                status_anterior=status_anterior,
                status_novo='aprovado',
                observacao="Aprovado no corte socioeconômico."
            )
            self.db.add(historico)
            
        self.db.commit()
        return len(aprovados)

    def matricular_candidato(self, cpf: str):
        """Efetiva a matrícula de um candidato aprovado, gerando o discente ativo"""
        cpf_limpo = ''.join(c for c in cpf if c.isdigit())
        cand = self.db.query(Candidato).filter(Candidato.cpf == cpf_limpo).first()
        
        if not cand:
            return False, "❌ Candidato não encontrado."
        if cand.status != 'aprovado':
            return False, f"⚠️ Candidato não está aprovado (Status: {cand.status})."
            
        # Muda status do candidato
        status_anterior = cand.status
        cand.status = 'confirmado'
        
        # Gera o aluno com cartão aleatório de 6 dígitos
        novo_cartao = secrets.randbelow(900000) + 100000
        novo_aluno = Aluno(cartao_id=novo_cartao, nome=cand.nome, status='ATIVADO')
        
        self.db.add(novo_aluno)
        
        # Registrar no histórico
        historico = HistoricoStatusCandidato(
            candidato_id=cand.id,
            candidato_nome=cand.nome,
            candidato_cpf=cand.cpf,
            status_anterior=status_anterior,
            status_novo='confirmado',
            observacao=f"Matrícula efetuada. Aluno gerado com Cartão ID: {novo_cartao}"
        )
        self.db.add(historico)
        
        self.db.commit()
        
        return True, f"✅ Matrícula confirmada! Aluno {cand.nome} gerado com Cartão ID: {novo_cartao}"

    def pesquisar_candidatos(self, termo: str):
        """Busca candidatos por nome ou CPF"""
        termo_limpo = termo.strip()
        termo_digits = ''.join(c for c in termo_limpo if c.isdigit())
        
        query = self.db.query(Candidato)
        if termo_digits:
            return query.filter(Candidato.cpf.like(f"%{termo_digits}%")).all()
        return query.filter(Candidato.nome.ilike(f"%{termo_limpo}%")).all()

    def excluir_candidato(self, candidato_id: int):
        """Remove fisicamente um candidato do banco de dados"""
        cand = self.db.query(Candidato).filter(Candidato.id == candidato_id).first()
        if not cand:
            return False, "❌ Candidato não encontrado."
            
        # Registrar exclusão no histórico (ID ficará nulo pela FK ondelete SET NULL, mas preservamos auditoria)
        historico = HistoricoStatusCandidato(
            candidato_id=None,
            candidato_nome=cand.nome,
            candidato_cpf=cand.cpf,
            status_anterior=cand.status,
            status_novo='excluido',
            observacao="Candidato excluído fisicamente do sistema."
        )
        self.db.add(historico)
        
        self.db.delete(cand)
        self.db.commit()
        return True, f"✅ Candidato {cand.nome} excluído com sucesso."

    def pesquisar_alunos(self, termo: str):
        """Busca alunos ativos ou desativados por nome ou ID do cartão"""
        termo_limpo = termo.strip()
        if termo_limpo.isdigit():
            return self.db.query(Aluno).filter(Aluno.cartao_id == int(termo_limpo)).all()
        return self.db.query(Aluno).filter(Aluno.nome.ilike(f"%{termo_limpo}%")).all()

    def desativar_aluno(self, aluno_id: int):
        """Exclui logicamente (desativa) um aluno do controle de presenças"""
        aluno = self.db.query(Aluno).filter(Aluno.id == aluno_id).first()
        if not aluno:
            return False, "❌ Aluno não encontrado."
        aluno.status = 'DESATIVADO'
        self.db.commit()
        return True, f"✅ Aluno {aluno.nome} desativado com sucesso!"

    def importar_candidatos_planilha(self, caminho_xlsx: str):
        """
        Importa candidatos a partir de uma planilha de respostas do Google Forms.
        Estrutura de colunas esperada (baseada no legado):
        - Nome: Coluna E (5)
        - CPF: Coluna I (9)
        - Email: Coluna D (4)
        - Respostas do questionário socioeconômico: Colunas K (11) a AI (35) -> Questões 10 a 34 (25 questões)
        """
        import openpyxl
        wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
        sheet = wb.active

        inseridos = 0
        erros = 0

        # Lendo os textos das perguntas (linha 1)
        perguntas_texto = {}
        for col_idx in range(11, 36):
            q_num = col_idx - 1  # K(11) -> q=10
            val = sheet.cell(row=1, column=col_idx).value
            perguntas_texto[q_num] = str(val).strip() if val is not None else f"Questão {q_num}"

        for row in range(2, sheet.max_row + 1):
            nome = sheet.cell(row=row, column=5).value
            email = sheet.cell(row=row, column=4).value
            cpf = sheet.cell(row=row, column=9).value

            if not nome or not cpf:
                continue

            respostas = {}
            for col_idx in range(11, 36):
                q_num = col_idx - 1  # K(11) -> q=10
                val = sheet.cell(row=row, column=col_idx).value
                respostas[q_num] = str(val) if val is not None else ""

            # Converte CPF e email para string
            cpf_str = str(cpf).split('.')[0].strip()
            email_str = str(email).strip() if email else ""

            sucesso, _ = self.cadastrar_candidato(str(nome), cpf_str, email_str, respostas, perguntas_texto)
            if sucesso:
                inseridos += 1
            else:
                erros += 1

        wb.close()
        return inseridos, erros

    def obter_historico_candidato(self, candidato_id: int):
        """Busca o histórico de transições de status de um candidato específico"""
        return self.db.query(HistoricoStatusCandidato).filter(
            HistoricoStatusCandidato.candidato_id == candidato_id
        ).order_by(HistoricoStatusCandidato.data_alteracao.asc()).all()

    def obter_historico_geral(self):
        """Retorna todo o histórico de status de candidatos (incluindo deletados)"""
        return self.db.query(HistoricoStatusCandidato).order_by(
            HistoricoStatusCandidato.data_alteracao.desc()
        ).all()

    def sincronizar_importacao_drive(self):
        """
        Baixa automaticamente o arquivo de respostas do Google Drive
        e executa a importação dos candidatos.
        """
        temp_path = "temp_drive_import.xlsx"
        try:
            # Baixa do Google Drive
            GoogleDriveDownloader.baixar_planilha_forms(temp_path)
            
            # Importa os candidatos
            inseridos, erros = self.importar_candidatos_planilha(temp_path)
            return True, f"✅ Sincronização concluída! {inseridos} candidatos importados, {erros} erros/duplicados."
        except Exception as e:
            return False, f"❌ Erro na sincronização com o Google Drive: {e}"
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    def obter_respostas_candidato(self, candidato_id: int):
        """Retorna todas as respostas daquele candidato ordenadas pelo número da pergunta"""
        return self.db.query(PontuacaoQuestionario).filter(
            PontuacaoQuestionario.candidato_id == candidato_id
        ).order_by(PontuacaoQuestionario.pergunta_numero.asc()).all()

    def obter_questoes_disponiveis(self):
        """Retorna os enunciados de perguntas únicas já cadastradas na base"""
        resultados = self.db.query(
            PontuacaoQuestionario.pergunta_numero,
            PontuacaoQuestionario.questao
        ).distinct().order_by(PontuacaoQuestionario.pergunta_numero.asc()).all()
        return [{"pergunta_numero": r[0], "questao": r[1]} for r in resultados if r[1]]

    def obter_estatisticas_questao(self, questao_texto: str):
        """Retorna a distribuição estatística de respostas para uma determinada questão"""
        from sqlalchemy import func
        total = self.db.query(PontuacaoQuestionario).filter(
            PontuacaoQuestionario.questao == questao_texto
        ).count()

        if total == 0:
            return []

        agrupado = self.db.query(
            PontuacaoQuestionario.resposta,
            func.count(PontuacaoQuestionario.id).label('quantidade')
        ).filter(
            PontuacaoQuestionario.questao == questao_texto
        ).group_by(
            PontuacaoQuestionario.resposta
        ).order_by(
            func.count(PontuacaoQuestionario.id).desc()
        ).all()

        estatisticas = []
        for r_texto, qtd in agrupado:
            pct = round((qtd / total) * 100, 2)
            estatisticas.append({
                "resposta": r_texto if r_texto else "Sem resposta",
                "quantidade": qtd,
                "percentual": pct
            })
        return estatisticas

    def filtrar_candidatos_por_resposta(self, questao_texto: str, resposta_texto: str = None):
        """Retorna a lista de candidatos (com suas respostas) para uma determinada questão, opcionalmente filtrando por uma resposta específica"""
        query = self.db.query(
            Candidato.id,
            Candidato.nome,
            Candidato.cpf,
            Candidato.email,
            Candidato.status,
            Candidato.pontuacao_socioeconomica,
            PontuacaoQuestionario.resposta
        ).join(
            PontuacaoQuestionario, Candidato.id == PontuacaoQuestionario.candidato_id
        ).filter(
            PontuacaoQuestionario.questao == questao_texto
        )
        
        if resposta_texto is not None:
            query = query.filter(PontuacaoQuestionario.resposta == resposta_texto)
            
        resultados = query.all()
        return [{
            "id": r[0],
            "nome": r[1],
            "cpf": r[2],
            "email": r[3],
            "status": r[4],
            "pontuacao_socioeconomica": r[5],
            "resposta": r[6]
        } for r in resultados]

    def obter_dados_exportacao(self, questoes_numeros: List[int]):
        """Retorna colunas e linhas para exportação com base nos números de questões selecionados."""
        # 1. Obter os enunciados das questões correspondentes aos números
        from ..database.models import Candidato, PontuacaoQuestionario
        
        questoes_db = self.db.query(
            PontuacaoQuestionario.pergunta_numero,
            PontuacaoQuestionario.questao
        ).filter(
            PontuacaoQuestionario.pergunta_numero.in_(questoes_numeros)
        ).distinct().all()
        
        # Mapeia pergunta_numero -> questao enunciado
        mapa_questoes = {q[0]: q[1] for q in questoes_db if q[0] is not None}
        
        # Colunas da planilha: Nome sempre em primeiro, depois as questões na ordem selecionada
        colunas = ["Nome"]
        for q_num in questoes_numeros:
            colunas.append(mapa_questoes.get(q_num, f"Questão {q_num}"))
            
        # 2. Obter todos os candidatos
        candidatos = self.db.query(Candidato).order_by(Candidato.nome.asc()).all()
        
        linhas = []
        for cand in candidatos:
            linha = [cand.nome.upper()]
            # Obter respostas deste candidato
            respostas_cand = self.db.query(
                PontuacaoQuestionario.pergunta_numero,
                PontuacaoQuestionario.resposta
            ).filter(
                PontuacaoQuestionario.candidato_id == cand.id,
                PontuacaoQuestionario.pergunta_numero.in_(questoes_numeros)
            ).all()
            
            mapa_respostas = {r[0]: r[1] for r in respostas_cand if r[0] is not None}
            for q_num in questoes_numeros:
                linha.append(mapa_respostas.get(q_num, ""))
            linhas.append(linha)
            
        return colunas, linhas

    def atualizar_presencas_leitor_service(self):
        """Sincroniza presenças com a planilha presenca.xlsx (Drive ou local) e atualiza frequência dos alunos"""
        import openpyxl
        from ..utils.google_drive import GoogleDriveDownloader
        from googleapiclient.http import MediaFileUpload
        import random
        
        caminho_local_presenca = "Entradas/presenca.xlsx"
        os.makedirs("Entradas", exist_ok=True)
        os.makedirs("Cartola mágica", exist_ok=True)
        
        service = None
        drive_file_id = None
        usando_drive = False
        
        try:
            service = GoogleDriveDownloader.obter_servico()
            # Busca 'presenca.xlsx' na pasta do Drive
            folder_id = '1XTtbqEyBY_gUr1ub3srXugpQCD4t0Yhj'
            query = f"name='presenca.xlsx' and '{folder_id}' in parents"
            response = service.files().list(q=query, spaces='drive', fields='files(id)').execute()
            files = response.get('files', [])
            if files:
                drive_file_id = files[0]['id']
                request = service.files().get_media(fileId=drive_file_id)
                with open(caminho_local_presenca, 'wb') as fh:
                    from googleapiclient.http import MediaIoBaseDownload
                    downloader = MediaIoBaseDownload(fh, request)
                    done = False
                    while not done:
                        status, done = downloader.next_chunk()
                usando_drive = True
                print("Download de presenca.xlsx do Google Drive realizado com sucesso.")
        except Exception as e:
            print(f"Aviso: Não foi possível conectar ao Google Drive ({e}). Utilizando arquivo local em: {caminho_local_presenca}")
            
        if not os.path.exists(caminho_local_presenca):
            # Se não existir nem localmente, cria uma planilha básica de presenças com 200 cartões vazios
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Presencas"
            ws.cell(row=1, column=1, value="Cartao")
            ws.cell(row=1, column=2, value="Status")
            ws.cell(row=1, column=3, value="Presencas")
            ws.cell(row=1, column=4, value="Aulas")
            ws.cell(row=1, column=5, value="Aluno")
            for i in range(1, 201):
                ws.cell(row=i+1, column=1, value=i)
                ws.cell(row=i+1, column=2, value="DESATIVADO")
                ws.cell(row=i+1, column=3, value=0)
                ws.cell(row=i+1, column=4, value=0)
                ws.cell(row=i+1, column=5, value="-")
            wb.save(caminho_local_presenca)
            wb.close()
            
        # 2. Carrega a planilha
        wb = openpyxl.load_workbook(caminho_local_presenca, data_only=True)
        ws = wb.active
        
        # 3. Mapeia cartões ativos no banco de dados e distribui para quem não tem
        alunos_ativos = self.db.query(Aluno).filter(Aluno.status == 'ATIVADO').all()
        
        # Coleta cartões já em uso
        cartoes_usados = {a.cartao_id for a in alunos_ativos if a.cartao_id is not None}
        # Cartões disponíveis (1 a 200)
        cartoes_livres = set(range(1, 201)) - cartoes_usados
        
        logs = []
        
        # Atribui cartões
        for aluno in alunos_ativos:
            if aluno.cartao_id is None or aluno.cartao_id == 0:
                if not cartoes_livres:
                    logs.append(f"⚠️ Sem cartões livres para atribuir a {aluno.nome}")
                    continue
                novo_c = random.choice(list(cartoes_livres))
                cartoes_livres.discard(novo_c)
                aluno.cartao_id = novo_c
                logs.append(f"💙 Cartão {novo_c} concedido a {aluno.nome} - e-mail notificação enviado (simulado).")
                
        # Mapeia aluno por cartao_id
        mapa_alunos = {a.cartao_id: a for a in alunos_ativos if a.cartao_id is not None}
        
        # 4. Atualiza a planilha e as frequências no banco de dados
        # Varre as linhas da planilha de presenças (linha 2 em diante)
        for r in range(2, ws.max_row + 1):
            c_val = ws.cell(row=r, column=1).value
            if c_val is None:
                continue
            try:
                c_id = int(c_val)
            except ValueError:
                continue
                
            if c_id in mapa_alunos:
                aluno = mapa_alunos[c_id]
                
                pres_val = ws.cell(row=r, column=3).value or 0
                aulas_val = ws.cell(row=r, column=4).value or 0
                
                try:
                    presencas_novas = float(pres_val)
                    aulas_novas = float(aulas_val)
                except (ValueError, TypeError):
                    presencas_novas = 0.0
                    aulas_novas = 0.0
                    
                carga_ant = float(aluno.carga_horaria_total or 0.0)
                perc_ant = float(aluno.percentual_presenca or 0.0)
                
                if carga_ant + aulas_novas != 0:
                    novo_perc = (carga_ant * perc_ant + aulas_novas * presencas_novas) / (carga_ant + aulas_novas)
                    aluno.percentual_presenca = round(novo_perc, 2)
                else:
                    aluno.percentual_presenca = 0.0
                    
                aluno.carga_horaria_total = round(carga_ant + aulas_novas, 3)
                
                # Zera counters na planilha
                ws.cell(row=r, column=2, value="ATIVADO")
                ws.cell(row=r, column=3, value=0)
                ws.cell(row=r, column=4, value=0)
                ws.cell(row=r, column=5, value=aluno.nome.upper())
                
                logs.append(f"💚 Cartão {c_id} ATIVADO | Presença de {aluno.nome} atualizada.")
            else:
                # Desativa cartão se não estiver associado a nenhum aluno
                ws.cell(row=r, column=2, value="DESATIVADO")
                ws.cell(row=r, column=3, value=0)
                ws.cell(row=r, column=4, value=0)
                ws.cell(row=r, column=5, value="-")
                if c_val in cartoes_usados:
                    logs.append(f"❤️ Cartão {c_id} DESATIVADO")
                    
        self.db.commit()
        
        # 5. Salva a planilha localmente
        caminho_salvar_presenca = "Cartola mágica/presenca.xlsx"
        wb.save(caminho_salvar_presenca)
        wb.close()
        
        # Copia também para Entradas/presenca.xlsx para manter atualizado
        openpyxl.load_workbook(caminho_salvar_presenca).save(caminho_local_presenca)
        
        # 6. Upload de volta para o Google Drive se estiver usando
        if usando_drive and service and drive_file_id:
            try:
                service.files().delete(fileId=drive_file_id).execute()
                file_metadata = {
                    'name': 'presenca.xlsx',
                    'parents': ['1XTtbqEyBY_gUr1ub3srXugpQCD4t0Yhj']
                }
                media = MediaFileUpload(caminho_salvar_presenca, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                service.files().create(body=file_metadata, media_body=media, supportsAllDrives=True).execute()
                logs.append("Upload da planilha mensageira de presenças concluído com sucesso.")
            except Exception as e:
                logs.append(f"⚠️ Erro ao enviar planilha atualizada para o Drive: {e}")
                
        return logs

    def modificar_dados_presenca_service(self, cartoes: List[int], horas: float, tipo: str):
        """Modifica as horas de presença dos alunos selecionados pelos cartões."""
        from ..database.models import Aluno
        alunos = self.db.query(Aluno).filter(Aluno.cartao_id.in_(cartoes), Aluno.status == 'ATIVADO').all()
        
        acertos = []
        erros = list(set(cartoes) - {a.cartao_id for a in alunos})
        
        for aluno in alunos:
            hant = (float(aluno.percentual_presenca or 0.0) * float(aluno.carga_horaria_total or 0.0)) / 100.0
            htotal = float(aluno.carga_horaria_total or 0.0)
            htotal0 = htotal
            
            if tipo == 'e':
                hmod = max(0.0, min(htotal, hant + horas))
            else:
                hmod = max(0.0, hant + horas)
                htotal = max(0.0, htotal + horas)
                
            if htotal != 0.0:
                aluno.percentual_presenca = round((hmod / htotal) * 100.0, 2)
            else:
                aluno.percentual_presenca = 0.0
                
            aluno.carga_horaria_total = round(htotal, 3)
            
            acertos.append({
                "nome": aluno.nome,
                "cartao": aluno.cartao_id,
                "hant": hant,
                "htotal0": htotal0,
                "pant": round(hant * 100 / htotal0) if htotal0 != 0 else 0,
                "hmod": hmod,
                "htotal": htotal,
                "pmod": round(aluno.percentual_presenca),
                "variacao": round(hmod - hant if abs(hmod - hant) >= abs(htotal - htotal0) else htotal - htotal0, 1)
            })
            
        self.db.commit()
        return acertos, erros
