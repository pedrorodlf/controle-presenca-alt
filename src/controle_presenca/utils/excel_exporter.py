import openpyxl
from typing import List, Any

class ExcelExporter:
    @staticmethod
    def gerar_planilha(colunas: List[str], linhas: List[List[Any]], caminho_salvar: str):
        """Gera uma planilha Excel com as colunas e linhas especificadas e salva no caminho indicado."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lista de Dados"
        
        # Escreve o cabeçalho
        for col_idx, col_name in enumerate(colunas, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            
        # Escreve as linhas
        for row_idx, row_data in enumerate(linhas, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
                
        wb.save(caminho_salvar)
        wb.close()
