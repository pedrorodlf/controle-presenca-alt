import os
import openpyxl
from typing import Dict, Any

class ScoreCalculator:
    _criterios = None

    @classmethod
    def _carregar_criterios(cls) -> Dict[int, Dict[str, float]]:
        if cls._criterios is not None:
            return cls._criterios

        # Caminho padrão para a planilha de critérios
        diretorio_atual = os.path.dirname(os.path.dirname(__file__))
        caminho_xlsx = os.path.join(diretorio_atual, 'config', 'criterios_pontuacao.xlsx')
        
        if not os.path.exists(caminho_xlsx):
            # Se não existir (ex: em ambiente de teste), retorna dicionário vazio
            return {}

        wb = openpyxl.load_workbook(caminho_xlsx, data_only=True)
        sheet = wb.active

        cls._criterios = {}
        current_q = None

        # Lê a partir da linha 3
        for row in range(3, sheet.max_row + 1):
            q_val = sheet.cell(row=row, column=1).value
            if q_val is not None:
                try:
                    current_q = int(q_val)
                except ValueError:
                    current_q = q_val
                cls._criterios[current_q] = {}

            alt_val = sheet.cell(row=row, column=2).value
            pts_val = sheet.cell(row=row, column=3).value

            if current_q is not None and alt_val is not None:
                alt_clean = str(alt_val).strip()
                try:
                    pts = float(pts_val) if pts_val is not None else 0.0
                except (ValueError, TypeError):
                    pts = 0.0
                cls._criterios[current_q][alt_clean] = pts

        wb.close()
        return cls._criterios

    @classmethod
    def calcular_score(cls, respostas: Dict[int, str]) -> float:
        """
        Calcula a pontuação socioeconômica total com base em um dicionário
        de respostas do candidato, mapeando {numero_questao: alternativa_escolhida}.
        """
        detalhes = cls.calcular_detalhes(respostas)
        return sum(d["pontos"] for d in detalhes.values())

    @classmethod
    def calcular_detalhes(cls, respostas: Dict[int, str]) -> Dict[int, Dict[str, Any]]:
        """
        Calcula a pontuação de cada questão individualmente,
        retornando mapeamento {numero_questao: {"resposta": resposta_dada, "pontos": pontos_calculados}}.
        """
        criterios = cls._carregar_criterios()
        detalhes = {}

        for q_num, resposta in respostas.items():
            pts = 0.0
            resp_clean = str(resposta).strip() if resposta is not None else ""
            if q_num in criterios:
                if resp_clean in criterios[q_num]:
                    pts = criterios[q_num][resp_clean]
                else:
                    matched = False
                    for key, val in criterios[q_num].items():
                        if key.lower() == resp_clean.lower():
                            pts = val
                            matched = True
                            break
            detalhes[q_num] = {"resposta": resp_clean, "pontos": pts}

        return detalhes
